import streamlit as st
import pandas as pd
import re
import pytesseract
from pdf2image import convert_from_path
from openpyxl import Workbook, load_workbook
from PIL import Image, ImageEnhance, ImageFilter, ImageOps
import os
import io

# =========================================================
# การตั้งค่า Tesseract Path สำหรับ Cloud Server
# =========================================================
try:
    # กำหนด Tesseract Path สำหรับ Linux/Cloud Server
    pytesseract.pytesseract.tesseract_cmd = '/usr/bin/tesseract'
except Exception:
    pass

def advanced_image_preprocessing(image):
    """ปรับปรุงรูปภาพสำหรับ OCR ด้วย PIL เท่านั้น (ไม่ใช้ OpenCV)"""
    try:
        # แปลงเป็น grayscale
        if image.mode != 'L':
            gray_image = image.convert('L')
        else:
            gray_image = image
        
        # 1. เพิ่ม contrast
        enhancer = ImageEnhance.Contrast(gray_image)
        contrast_image = enhancer.enhance(2.5)
        
        # 2. เพิ่ม sharpness
        enhancer = ImageEnhance.Sharpness(contrast_image)
        sharp_image = enhancer.enhance(2.0)
        
        # 3. ปรับ brightness
        enhancer = ImageEnhance.Brightness(sharp_image)
        bright_image = enhancer.enhance(1.2)
        
        # 4. Apply filters
        # Sharpen filter
        sharpened = bright_image.filter(ImageFilter.SHARPEN)
        
        # Edge enhance
        edge_enhanced = sharpened.filter(ImageFilter.EDGE_ENHANCE_MORE)
        
        # 5. Auto contrast
        final_image = ImageOps.autocontrast(edge_enhanced, cutoff=2)
        
        # 6. เพิ่ม contrast อีกครั้ง
        enhancer = ImageEnhance.Contrast(final_image)
        final_processed = enhancer.enhance(1.3)
        
        return final_processed
        
    except Exception as e:
        st.warning(f"⚠️ ไม่สามารถประมวลผลรูปภาพขั้นสูงได้: {e}")
        # Fallback เป็นการปรับปรุงพื้นฐาน
        return enhance_image_basic(image)

def enhance_image_basic(image):
    """การปรับปรุงรูปภาพพื้นฐาน (Fallback)"""
    enhancer = ImageEnhance.Contrast(image)
    image = enhancer.enhance(2.0)
    
    enhancer = ImageEnhance.Sharpness(image)
    image = enhancer.enhance(1.5)
    
    enhancer = ImageEnhance.Brightness(image)
    image = enhancer.enhance(1.1)
    
    return image

def extract_ocr_from_pdf(pdf_bytes):
    """แปลง PDF เป็น OCR Text และคืนค่าทั้ง text และ images"""
    temp_file = "temp_upload.pdf"
    try:
        # บันทึกไฟล์ชั่วคราว
        with open(temp_file, "wb") as f:
            f.write(pdf_bytes)
        
        st.info("🔄 กำลังแปลง PDF เป็นรูปภาพ...")
        
        # แปลง PDF เป็นรูปภาพด้วยความละเอียดสูง
        pages = convert_from_path(temp_file, dpi=450, fmt='PNG')
        
        ocr_results = []
        
        for i, page in enumerate(pages):
            st.info(f"📖 กำลังทำ OCR หน้าที่ {i+1}/{len(pages)}...")
            
            # ใช้การประมวลผลรูปภาพขั้นสูง
            enhanced_page = advanced_image_preprocessing(page)
            
            # ลอง OCR หลายแบบเพื่อเพิ่มความแม่นยำ
            ocr_configs = [
                '--psm 6 --oem 3',  # Uniform text block
                '--psm 4 --oem 3',  # Single column text
                '--psm 8 --oem 3',  # Single word
                '--psm 13 --oem 3'  # Raw line. Treat the image as a single text line
            ]
            
            best_text = ""
            best_confidence = 0
            
            for config in ocr_configs:
                try:
                    # ทำ OCR ด้วย config นี้
                    ocr_data = pytesseract.image_to_data(
                        enhanced_page,
                        lang="tha+eng",
                        config=config,
                        output_type=pytesseract.Output.DICT
                    )
                    
                    # คำนวณ confidence เฉลี่ย
                    confidences = [int(conf) for conf in ocr_data['conf'] if int(conf) > 0]
                    avg_confidence = sum(confidences) / len(confidences) if confidences else 0
                    
                    # เลือก result ที่มี confidence สูงสุด
                    if avg_confidence > best_confidence:
                        best_confidence = avg_confidence
                        best_text = pytesseract.image_to_string(
                            enhanced_page,
                            lang="tha+eng",
                            config=config
                        )
                        
                except Exception as e:
                    continue
            
            # ถ้าไม่มี result ที่ดี ให้ใช้การ OCR พื้นฐาน
            if not best_text.strip():
                best_text = pytesseract.image_to_string(
                    enhanced_page,
                    lang="tha+eng",
                    config='--psm 6 --oem 3'
                )
            
            ocr_results.append({
                'page_number': i + 1,
                'ocr_text': best_text,
                'image': enhanced_page,
                'confidence': best_confidence
            })
        
        # ลบไฟล์ชั่วคราว
        os.remove(temp_file)
        
        return ocr_results
        
    except Exception as e:
        if os.path.exists(temp_file):
            os.remove(temp_file)
        st.error(f"❌ เกิดข้อผิดพลาดในการแปลง PDF: {str(e)}")
        return []

def clean_amount(raw_amount):
    """ทำความสะอาดตัวเลขที่ดึงมา"""
    if not raw_amount:
        return ""
    
    # ลบเครื่องหมายและอักขระพิเศษ
    cleaned = re.sub(r'[^\d\.]', '', raw_amount.replace(',', ''))
    
    try:
        # ตรวจสอบว่าเป็นตัวเลขที่สมเหตุสมผล
        float_val = float(cleaned)
        if 1 <= float_val <= 999999999:  # ช่วงที่เหมาะสม
            return f"{float_val:.2f}"
        return ""
    except (ValueError, TypeError):
        return ""

def extract_data_from_ocr_text(text):
    """ดึงข้อมูลจากข้อความ OCR ด้วยความแม่นยำสูง"""
    data = {
        'date': '',
        'invoice_number': '',
        'amount': '',
        'raw_matches': {},
        'debug_info': {}
    }
    
    # ทำความสะอาดข้อความก่อน
    clean_text = re.sub(r'\s+', ' ', text.strip())
    
    # === 1. ดึงเลขที่เอกสาร (HH Pattern) ===
    invoice_patterns = [
        r'(?:เลขที[่ิ]*|No\.?|Invoice\s*No\.?)[:\s]*([HH]{1,2}\d{6,8})',  # มี label นำหน้า
        r'\b(HH\d{6,8})\b',  # HH ตามด้วยตัวเลข 6-8 หลัก
        r'([H]{2}\d{6,8})',  # HH แล้วตามด้วยตัวเลข
        r'(HH[\s]*\d{6,8})',  # HH อาจมีช่องว่างคั่น
    ]
    
    for pattern in invoice_patterns:
        matches = re.findall(pattern, clean_text, re.IGNORECASE)
        if matches:
            # เลือกเลขที่ยาวที่สุดและสมเหตุสมผล
            valid_invoices = [m for m in matches if len(re.sub(r'[^A-Z0-9]', '', m)) >= 8]
            if valid_invoices:
                data['invoice_number'] = valid_invoices[0].replace(' ', '')
                data['raw_matches']['invoices_found'] = matches
                break
    
    # === 2. ดึงวันที่ (Date Pattern) ===
    date_patterns = [
        r'(?:วันที[่ิ]*|Date)[:\s]*(\d{1,2}/\d{1,2}/\d{2,4})',  # มี label
        r'\b(\d{1,2}/\d{1,2}/\d{2})\b',  # รูปแบบ DD/MM/YY
        r'(\d{2}/\d{2}/\d{2})',  # รูปแบบ DD/MM/YY แน่นอน
        r'(\d{1,2}[-/]\d{1,2}[-/]\d{2,4})',  # รองรับทั้ง / และ -
    ]
    
    for pattern in date_patterns:
        matches = re.findall(pattern, clean_text, re.IGNORECASE)
        if matches:
            # เลือกวันที่ที่มีรูปแบบถูกต้อง
            valid_dates = []
            for date_str in matches:
                if re.match(r'\d{1,2}[/-]\d{1,2}[/-]\d{2,4}$', date_str):
                    # แปลง - เป็น /
                    normalized_date = date_str.replace('-', '/')
                    valid_dates.append(normalized_date)
            
            if valid_dates:
                data['date'] = valid_dates[0]
                data['raw_matches']['dates_found'] = matches
                break
    
    # === 3. ดึงยอดเงิน (Amount Pattern) - ขั้นสูงขึ้น ===
    amount_patterns = [
        # Pattern 1: หาจากคำว่า "มูลค่าสินค้า" หรือ "Product Value"
        r'(?:มูลค[่า]*สินค[้า]*|Product\s*Value)[:\s]*([,\d]+\.?\d{0,2})',
        
        # Pattern 2: หาตัวเลขที่อยู่ในบรรทัดที่มีคำว่า "มูลค่า"
        r'มูลค[่า]*[^0-9\n]*([,\d]+\.\d{2})',
        
        # Pattern 3: หาตัวเลขที่อยู่ก่อน "จำนวนภาษี" หรือ "VAT"
        r'([,\d]+\.\d{2})\s*(?:จำนวนภาษี|VAT|7\.00\s*%)',
        
        # Pattern 4: หาตัวเลขที่อยู่หลัง "หักส่วนลด" และก่อน "VAT"
        r'(?:หักส่วนลด|Discount)[\s\S]*?([,\d]+\.\d{2})[\s\S]*?(?:VAT|ภาษี)',
        
        # Pattern 5: หาตัวเลขทศนิยม 2 ตำแหน่งที่มีขนาดเหมาะสม
        r'\b([,\d]{4,}\.\d{2})\b',
        
        # Pattern 6: หาในบริบท subtotal หรือ net amount
        r'(?:รวม|Total|Net|Sub)[^0-9]*([,\d]+\.\d{2})',
    ]
    
    found_amounts = []
    
    for i, pattern in enumerate(amount_patterns):
        matches = re.findall(pattern, clean_text, re.IGNORECASE | re.DOTALL)
        for match in matches:
            cleaned_amount = clean_amount(match)
            if cleaned_amount:
                # เก็บข้อมูลเพิ่มเติมว่าได้มาจาก pattern ไหน
                found_amounts.append({
                    'amount': cleaned_amount,
                    'raw': match,
                    'pattern': i + 1,
                    'numeric_value': float(cleaned_amount)
                })
    
    # เลือกยอดเงินที่เหมาะสมที่สุด
    if found_amounts:
        # เรียงตาม pattern priority และความเหมาะสมของตัวเลข
        found_amounts.sort(key=lambda x: (x['pattern'], -x['numeric_value']))
        
        # เลือกจากที่มี pattern ดีที่สุด หรือตัวเลขที่สมเหตุสมผล
        best_amount = found_amounts[0]
        
        # ตรวจสอบความสมเหตุสมผลเพิ่มเติม
        if 100 <= best_amount['numeric_value'] <= 100000:  # ช่วงที่เหมาะสมสำหรับใบเสร็จ
            data['amount'] = best_amount['amount']
        
        data['raw_matches']['amounts_found'] = [amt['raw'] for amt in found_amounts[:5]]  # เก็บ 5 อันแรก
        data['debug_info']['amount_details'] = found_amounts[:3]  # เก็บรายละเอียดเพื่อ debug
    
    return data

def create_excel_template():
    """สร้างไฟล์ Excel Template"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice_Data"
    
    # Header
    headers = ['ลำดับ', 'วันที่', 'เลขที่ตามบิล', 'ยอดก่อน VAT']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # สร้างไฟล์ในหน่วยความจำ
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def fill_excel_with_data(data_list):
    """กรอกข้อมูลลง Excel"""
    # สร้าง DataFrame จาก List of Dictionaries
    df_data = pd.DataFrame(data_list)
    
    # จัดเรียงตามเลขหน้า
    df_data = df_data.sort_values(by='page_number').reset_index(drop=True)
    
    # แปลงเป็น Excel
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # เลือกเฉพาะคอลัมน์ที่ต้องการ
        df_to_excel = df_data[['date', 'invoice_number', 'amount']].copy()
        df_to_excel.insert(0, 'ลำดับ', df_to_excel.index + 1)
        
        # เปลี่ยนชื่อคอลัมน์ให้ตรงกับที่ต้องการ
        df_to_excel.columns = ['ลำดับ', 'วันที่', 'เลขที่ตามบิล', 'ยอดก่อน VAT']
        
        df_to_excel.to_excel(writer, index=False, sheet_name='Invoice_Data')
    
    output.seek(0)
    return output

def main():
    st.set_page_config(
        page_title="Enhanced PDF OCR Extractor",
        layout="wide",
        initial_sidebar_state="expanded"
    )
    
    st.title("🔍 Enhanced PDF OCR Extractor - High Accuracy")
    st.markdown("**ปรับปรุงความแม่นยำสำหรับใบเสร็จภาษาไทย**")
    st.markdown("---")
    
    # Initialize session state
    if 'ocr_results' not in st.session_state:
        st.session_state.ocr_results = []
    if 'extracted_data' not in st.session_state:
        st.session_state.extracted_data = []
    
    # Sidebar
    with st.sidebar:
        st.header("⚙️ ขั้นตอนการทำงาน")
        st.markdown("""
        1. 📁 **อัปโหลด PDF**
        2. 🚀 **กด 'เริ่ม OCR ขั้นสูง'**
        3. ✏️ **ตรวจสอบ/แก้ไข**
        4. 💾 **ดาวน์โหลด Excel**
        """)
        
        st.markdown("---")
        st.markdown("### 🎯 การปรับปรุงใหม่:")
        st.markdown("""
        ✅ **PIL Image Processing**
        - Auto Contrast Enhancement
        - Edge Detection & Sharpening
        - Brightness & Contrast Optimization
        - Multi-layer Filtering
        
        ✅ **Multi-Pattern Recognition**
        - หลาย OCR Config
        - Confidence Scoring
        - Pattern Priority
        - Smart Validation
        """)
        
        st.markdown("### 📊 เป้าหมายข้อมูล:")
        st.markdown("""
        - **วันที่:** DD/MM/YY
        - **เลขที่:** HHxxxxxxx
        - **มูลค่า:** xxxx.xx
        """)
    
    # === Step 1: Upload PDF ===
    st.header("1. 📁 อัปโหลดไฟล์ PDF")
    uploaded_file = st.file_uploader(
        "เลือกไฟล์ PDF",
        type="pdf",
        help="อัปโหลดไฟล์ใบเสร็จ PDF (รองรับหลายหน้า)"
    )
    
    if uploaded_file is not None:
        col1, col2 = st.columns([2, 3])
        with col1:
            st.success(f"✅ ไฟล์: {uploaded_file.name}")
            file_size = len(uploaded_file.getvalue()) / (1024 * 1024)
            st.info(f"📊 ขนาด: {file_size:.1f} MB")
        
        with col2:
            if st.button("🚀 เริ่มแปลง OCR ขั้นสูง", type="primary", use_container_width=True):
                with st.spinner("⚡ กำลังประมวลผล PDF ด้วย AI OCR..."):
                    pdf_bytes = uploaded_file.getvalue()
                    st.session_state.ocr_results = extract_ocr_from_pdf(pdf_bytes)
        
        # === Step 2: Display OCR Results ===
        if st.session_state.ocr_results:
            st.success(f"✅ OCR เสร็จสิ้น! ประมวลผล {len(st.session_state.ocr_results)} หน้า")
            
            # แสดงสถิติ confidence
            avg_confidence = sum([r.get('confidence', 0) for r in st.session_state.ocr_results]) / len(st.session_state.ocr_results)
            st.info(f"📊 ความแม่นยำเฉลี่ย: {avg_confidence:.1f}%")
            
            st.header("2. 👁️ ตรวจสอบและแก้ไขข้อมูล")
            st.warning("⚠️ โปรดตรวจสอบข้อมูลในแต่ละหน้าและกด 'บันทึก'")
            
            # === แสดงผลลัพธ์แต่ละหน้า ===
            for result in st.session_state.ocr_results:
                page_key = result['page_number']
                confidence = result.get('confidence', 0)
                
                # ดึงข้อมูลจาก OCR
                extracted = extract_data_from_ocr_text(result['ocr_text'])
                
                # ค้นหาข้อมูลที่บันทึกไว้แล้ว
                saved_data = next((d for d in st.session_state.extracted_data if d['page_number'] == page_key), None)
                
                # กำหนดสีของ expander ตามสถานะ
                status_icon = "💾 บันทึกแล้ว" if saved_data else "✏️ รอการตรวจสอบ"
                confidence_badge = f"🎯 {confidence:.1f}%" if confidence > 0 else ""
                
                with st.expander(f"📄 หน้าที่ {page_key} | {status_icon} | {confidence_badge}", expanded=not saved_data):
                    
                    col1, col2 = st.columns([1, 1])
                    
                    with col1:
                        st.subheader("📝 ข้อมูลที่ตรวจพบ:")
                        
                        # ใช้ค่าที่บันทึกไว้หรือค่าจาก OCR ใหม่
                        initial_date = saved_data['date'] if saved_data else extracted['date']
                        initial_invoice = saved_data['invoice_number'] if saved_data else extracted['invoice_number']
                        initial_amount = saved_data['amount'] if saved_data else extracted['amount']
                        
                        # === Input Fields ===
                        date_value = st.text_input(
                            "📅 วันที่:",
                            value=initial_date,
                            help="รูปแบบ: DD/MM/YY เช่น 01/08/68",
                            key=f"date_{page_key}"
                        )
                        
                        invoice_value = st.text_input(
                            "🔢 เลขที่ตามบิล:",
                            value=initial_invoice,
                            help="รูปแบบ: HHxxxxxxx เช่น HH6800470",
                            key=f"invoice_{page_key}"
                        )
                        
                        amount_value = st.text_input(
                            "💰 ยอดก่อน VAT:",
                            value=initial_amount,
                            help="รูปแบบ: xxxx.xx เช่น 4710.28",
                            key=f"amount_{page_key}"
                        )
                        
                        # === Save Button ===
                        if st.button(f"💾 บันทึกข้อมูลหน้า {page_key}", key=f"save_{page_key}", type="primary", use_container_width=True):
                            
                            # ตรวจสอบและอัปเดตข้อมูล
                            existing_index = next((i for i, data in enumerate(st.session_state.extracted_data) if data['page_number'] == page_key), None)
                            
                            page_data = {
                                'page_number': page_key,
                                'date': date_value.strip(),
                                'invoice_number': invoice_value.strip(),
                                'amount': amount_value.strip()
                            }
                            
                            if existing_index is not None:
                                st.session_state.extracted_data[existing_index] = page_data
                            else:
                                st.session_state.extracted_data.append(page_data)
                            
                            st.success(f"✅ บันทึกข้อมูลหน้า {page_key} เรียบร้อย!")
                            st.rerun()
                    
                    with col2:
                        st.subheader("🖼️ รูปภาพที่ปรับปรุงแล้ว:")
                        st.image(result['image'], use_container_width=True)
                        
                        # === Debug Information ===
                        with st.expander("🔍 ข้อมูล Debug & OCR Text"):
                            st.text_area(
                                "OCR Raw Text:",
                                result['ocr_text'][:1000] + "..." if len(result['ocr_text']) > 1000 else result['ocr_text'],
                                height=200,
                                key=f"ocr_debug_{page_key}"
                            )
                            
                            if extracted['debug_info']:
                                st.json(extracted['debug_info'])
                            
                            if extracted['raw_matches']:
                                st.write("**ข้อมูลทั้งหมดที่พบ:**")
                                st.json(extracted['raw_matches'])
            
            # === Step 3: Create Excel ===
            if st.session_state.extracted_data:
                st.header("3. 💾 สร้างและดาวน์โหลด Excel")
                
                # แสดงสรุปข้อมูล
                df_summary = pd.DataFrame(st.session_state.extracted_data)
                df_summary = df_summary.sort_values(by='page_number').reset_index(drop=True)
                df_display = df_summary[['page_number', 'date', 'invoice_number', 'amount']].copy()
                df_display.columns = ['หน้า', 'วันที่', 'เลขที่ตามบิล', 'ยอดก่อน VAT']
                
                st.subheader("📋 สรุปข้อมูลสุดท้าย:")
                st.dataframe(df_display, use_container_width=True, height=300)
                
                # สถิติ
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("จำนวนหน้าที่บันทึก", len(df_summary))
                with col2:
                    valid_amounts = len([d for d in st.session_state.extracted_data if d['amount']])
                    st.metric("หน้าที่มียอดเงิน", valid_amounts)
                with col3:
                    try:
                        total_amount = sum([float(d['amount']) for d in st.session_state.extracted_data if d['amount']])
                        st.metric("ยอดรวมทั้งหมด", f"{total_amount:,.2f}")
                    except:
                        st.metric("ยอดรวมทั้งหมด", "ไม่สามารถคำนวณได้")
                
                # === Download Button ===
                st.markdown("---")
                excel_file = fill_excel_with_data(st.session_state.extracted_data)
                
                st.download_button(
                    label="⬇️ ดาวน์โหลดไฟล์ Excel (Final)",
                    data=excel_file,
                    file_name=f"Enhanced_Invoice_Data_{uploaded_file.name.replace('.pdf', '')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    use_container_width=True
                )
        
    else:
        st.info("👆 กรุณาอัปโหลดไฟล์ PDF เพื่อเริ่มต้น")
        
        # แสดง Template Excel สำหรับดาวน์โหลด
        st.header("📋 หรือดาวน์โหลด Excel Template")
        
        col1, col2 = st.columns([3, 1])
        with col1:
            st.markdown("""
            **คุณสามารถ:**
            - ดาวน์โหลด Template Excel เปล่า
            - อัปโหลด PDF แล้วให้ AI ดึงข้อมูลอัตโนมัติ
            - ตรวจสอบและแก้ไขข้อมูลก่อนดาวน์โหลด
            """)
        
        with col2:
            template_file = create_excel_template()
            st.download_button(
                label="⬇️ ดาวน์โหลด Template",
                data=template_file,
                file_name="Invoice_Template.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

if __name__ == "__main__":
    main()
